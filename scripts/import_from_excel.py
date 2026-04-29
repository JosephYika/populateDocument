"""
Import client data from KG Construction's Client Directory Excel spreadsheet.

Usage:
    python scripts/import_from_excel.py Client_Directory.xlsx

Handles the spreadsheet's specific layout:
  - Row 1: blank, Row 2: company header, Row 3: column headers
  - Company group separator rows (single-cell, all-caps or indented names)
  - Managing Company column contains multiline data (name + address + phone)
  - "No Management Company" marker for individual owners
  - Billing Method determines send_directly_to_client

Re-runnable / idempotent — uses upserts on stable keys:
  - Companies:  matched by name (case-insensitive)
  - Managers:   matched by email if present, else by (company_id, name)
  - Clients:    matched by (address, unit)

Never deletes existing records — only inserts and updates.
"""

import sys
import os
import re
import argparse
from datetime import datetime, timezone

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

from openpyxl import load_workbook
from sqlalchemy import func
from models import Session, ManagementCompany, PropertyManager, Client

# Spreadsheet layout: row 1 blank, row 2 company header, row 3 column headers.
HEADER_ROW = 3
DATA_START_ROW = 4

# Zero-based column positions matching the Client Directory spreadsheet layout.
COL_INDEX = {
    'address':         0,
    'city':            1,
    'client_name':     2,
    'client_email':    3,
    'client_phone':    4,
    'company_block':   5,   # Multiline cell: company name, address, phone
    'manager_name':    6,
    'pm_email':        7,
    'pm_phone':        8,
    'billing':         9,   # "Direct to Client" or via management company
    'notes':          10,
}

# Sentinel values in the company column indicating no management company.
NO_COMPANY_MARKERS = {'no management company', '— no management company —'}


def cell_str(row, col_key):
    """Read a cell value as a stripped string, or None if empty."""
    idx = COL_INDEX.get(col_key)
    if idx is None or idx >= len(row):
        return None
    val = row[idx].value
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def is_group_separator(row):
    """Detect company-group separator rows (blank rows or standalone headers)."""
    vals = [(i, cell.value) for i, cell in enumerate(row)
            if cell.value is not None and str(cell.value).strip()]
    if len(vals) == 0:
        return True
    if len(vals) == 1 and vals[0][0] == 0:
        return True
    if len(vals) == 2 and vals[0][0] == 0 and vals[1][0] == COL_INDEX['company_block']:
        text = str(vals[1][1]).strip().lower()
        if any(m in text for m in NO_COMPANY_MARKERS):
            return True
    return False


def is_footer_row(row):
    """Detect the 'KG Construction Corp' footer that marks end of data."""
    val = cell_str(row, 'address')
    if val and 'KG Construction Corp' in val:
        return True
    return False


def parse_company_block(block_text):
    """Parse a multiline company cell into (name, address, phone).

    The cell format is: line 1 = company name, remaining lines = address
    parts or phone (detected by leading digit/parenthesis pattern).
    Returns (None, None, None) for "No Management Company" markers.
    """
    if not block_text:
        return None, None, None
    clean = block_text.replace('—', '-').replace('–', '-')
    for marker in NO_COMPANY_MARKERS:
        if marker in clean.lower():
            return None, None, None
    lines = [l.strip() for l in clean.split('\n') if l.strip()]
    if not lines:
        return None, None, None
    name = lines[0]
    address_parts = []
    phone = None
    for line in lines[1:]:
        if re.match(r'^[\(\d][\d\s\-\(\)\.ext]+$', line):
            phone = line
        else:
            address_parts.append(line)
    address = ', '.join(address_parts) if address_parts else None
    return name, address, phone


def parse_billing(billing_text):
    """Return True if the billing method indicates direct-to-client."""
    if not billing_text:
        return False
    first_line = billing_text.split('\n')[0].strip().lower()
    return 'direct to client' in first_line


def clean_email(email_str):
    """Extract a valid email from a cell that may contain extra text."""
    if not email_str:
        return None
    first_line = email_str.split('\n')[0].strip()
    first_line = first_line.strip()
    if '@' in first_line:
        return first_line
    return None


def parse_address_unit(raw_address):
    """Split "123 Main St, Apt 4B" into ("123 Main St", "Apt 4B").

    Returns (address, unit) where unit is None if no apartment/unit/suite
    suffix is detected.
    """
    if not raw_address:
        return None, None
    apt_patterns = [
        r',?\s*(Apt\.?\s+\S+)$',
        r',?\s*(Unit\s+\S+)$',
        r',?\s*(Suite\s+\S+)$',
        r',?\s*(PHT|PHS?\s*\w*)$',
        r',?\s*(#\S+)$',
    ]
    address = raw_address.strip()
    unit = None
    for pat in apt_patterns:
        m = re.search(pat, address, re.IGNORECASE)
        if m:
            unit = m.group(1).strip()
            address = address[:m.start()].strip().rstrip(',')
            break
    return address, unit


def upsert_company(session, name, address, phone, stats):
    """Find or create a company by name (case-insensitive). Updates fields if changed."""
    if not name:
        return None

    existing = session.query(ManagementCompany).filter(
        func.lower(ManagementCompany.name) == name.lower()
    ).first()

    if existing:
        updated = False
        if address and existing.address != address:
            existing.address = address
            updated = True
        if phone and existing.phone != phone:
            existing.phone = phone
            updated = True
        if updated:
            existing.updated_at = datetime.now(timezone.utc)
            stats['companies_updated'] += 1
        else:
            stats['companies_existing'] += 1
        return existing

    company = ManagementCompany(name=name, address=address, phone=phone)
    session.add(company)
    session.flush()
    stats['companies_created'] += 1
    return company


def upsert_manager(session, name, email, phone, company, stats):
    """Find or create a manager. Matches by email first, then by (company, name)."""
    if not name:
        return None

    existing = None
    if email:
        existing = session.query(PropertyManager).filter(
            func.lower(PropertyManager.email) == email.lower()
        ).first()

    if not existing and company:
        existing = session.query(PropertyManager).filter(
            PropertyManager.company_id == company.id,
            func.lower(PropertyManager.name) == name.lower()
        ).first()

    if existing:
        updated = False
        if email and existing.email != email:
            existing.email = email
            updated = True
        if phone and existing.phone != phone:
            existing.phone = phone
            updated = True
        if company and existing.company_id != company.id:
            existing.company_id = company.id
            updated = True
        if updated:
            existing.updated_at = datetime.now(timezone.utc)
            stats['managers_updated'] += 1
        else:
            stats['managers_existing'] += 1
        return existing

    manager = PropertyManager(
        name=name,
        email=email,
        phone=phone,
        company_id=company.id if company else None,
    )
    session.add(manager)
    session.flush()
    stats['managers_created'] += 1
    return manager


def upsert_client(session, data, company, manager, stats):
    """Find or create a client by (address, unit). Updates changed fields."""
    address = data['address']
    unit = data['unit']

    q = session.query(Client).filter(
        func.lower(Client.address) == address.lower()
    )
    if unit:
        q = q.filter(func.lower(Client.unit) == unit.lower())
    else:
        q = q.filter((Client.unit == None) | (Client.unit == ''))

    existing = q.first()

    if existing:
        updated = False
        fields = {
            'owner_name': data['client_name'],
            'default_company_id': company.id if company else None,
            'default_manager_id': manager.id if manager else None,
            'send_directly_to_client': data['send_directly'],
            'client_email': data['client_email'],
        }
        if data['notes']:
            fields['notes'] = data['notes']

        for attr, val in fields.items():
            if val is not None and getattr(existing, attr) != val:
                setattr(existing, attr, val)
                updated = True

        if updated:
            existing.updated_at = datetime.now(timezone.utc)
            stats['clients_updated'] += 1
        else:
            stats['clients_existing'] += 1
        return

    client = Client(
        address=address,
        unit=unit,
        owner_name=data['client_name'],
        default_company_id=company.id if company else None,
        default_manager_id=manager.id if manager else None,
        send_directly_to_client=data['send_directly'],
        client_email=data['client_email'],
        notes=data['notes'],
    )
    session.add(client)
    session.flush()
    stats['clients_created'] += 1


def run_import(file_path):
    """Main import loop: read rows, upsert companies/managers/clients, print summary."""
    print(f'Loading workbook: {file_path}')
    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active
    print(f'Sheet: "{sheet.title}" — {sheet.max_row} total rows')

    rows = list(sheet.iter_rows(min_row=DATA_START_ROW))
    if not rows:
        print('ERROR: No data rows found.')
        return

    stats = {
        'companies_created': 0, 'companies_updated': 0, 'companies_existing': 0,
        'managers_created': 0, 'managers_updated': 0, 'managers_existing': 0,
        'clients_created': 0, 'clients_updated': 0, 'clients_existing': 0,
        'rows_skipped': 0, 'rows_processed': 0, 'errors': 0,
    }

    session = Session()

    try:
        for row_idx, row in enumerate(rows, start=DATA_START_ROW):
            if is_footer_row(row):
                break

            if is_group_separator(row):
                stats['rows_skipped'] += 1
                continue

            stats['rows_processed'] += 1

            try:
                raw_address = cell_str(row, 'address')
                city = cell_str(row, 'city')
                if not raw_address:
                    stats['rows_skipped'] += 1
                    continue

                address, unit = parse_address_unit(raw_address)
                if city:
                    address = f'{address}, {city}'

                company_block = cell_str(row, 'company_block')
                co_name, co_address, co_phone = parse_company_block(company_block)
                company = upsert_company(session, co_name, co_address, co_phone, stats)

                manager_name = cell_str(row, 'manager_name')
                pm_email = clean_email(cell_str(row, 'pm_email'))
                pm_phone = cell_str(row, 'pm_phone')
                manager = upsert_manager(session, manager_name, pm_email, pm_phone, company, stats)

                billing = cell_str(row, 'billing')
                send_directly = parse_billing(billing)

                client_email = clean_email(cell_str(row, 'client_email'))

                data = {
                    'address': address,
                    'unit': unit,
                    'client_name': cell_str(row, 'client_name'),
                    'send_directly': send_directly,
                    'client_email': client_email,
                    'notes': cell_str(row, 'notes'),
                }

                upsert_client(session, data, company, manager, stats)

            except Exception as e:
                stats['errors'] += 1
                print(f'  ERROR on row {row_idx}: {e}')

        session.commit()
        print('\nImport complete!')

    except Exception as e:
        session.rollback()
        print(f'\nFATAL ERROR — rolled back: {e}')
        raise
    finally:
        session.close()
        wb.close()

    print(f'\n{"="*40}')
    print(f'  IMPORT SUMMARY')
    print(f'{"="*40}')
    print(f'  Rows processed:      {stats["rows_processed"]}')
    print(f'  Rows skipped:        {stats["rows_skipped"]}')
    print(f'  Errors:              {stats["errors"]}')
    print()
    print(f'  Companies created:   {stats["companies_created"]}')
    print(f'  Companies updated:   {stats["companies_updated"]}')
    print(f'  Companies existing:  {stats["companies_existing"]}')
    print()
    print(f'  Managers created:    {stats["managers_created"]}')
    print(f'  Managers updated:    {stats["managers_updated"]}')
    print(f'  Managers existing:   {stats["managers_existing"]}')
    print()
    print(f'  Clients created:     {stats["clients_created"]}')
    print(f'  Clients updated:     {stats["clients_updated"]}')
    print(f'  Clients existing:    {stats["clients_existing"]}')
    print(f'{"="*40}')


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Import KG Construction client data from Excel')
    parser.add_argument('file', help='Path to the Client Directory .xlsx file')
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f'ERROR: File not found: {args.file}')
        sys.exit(1)

    run_import(args.file)
